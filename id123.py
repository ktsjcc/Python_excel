import os,sys
import openpyxl
import pandas as pd
import datetime

##用来判断性别的函数
def SEX():
    for i in ID:
        if int(i[-2])%2==1:
            sex.append('男')
        else:
            sex.append('女')
            
##提取生日的函数
def birthday():
    for i in ID:
        bday.append(i[6:14])

##计算年龄
def age():
    date=datetime.date.today()
    year=int(date.__str__()[:4])
    
    for i in bday:
        b_year=int(i[:4])
        A_ge=year-b_year
        Age.append(str(A_ge))

##判断女性年龄大于50岁的，男性年龄大于60岁的。
def Judge():
    for i,b in zip(sex,Age):
        if i=='女' and int(b)>=50:
            result.append('是')
        elif i=='女' and int(b)<50:
            result.append('否')
        elif i=='男' and int(b)>=60:
            result.append('是')
        elif i=='男' and int(b)<60:
            result.append('否')
            
            

#写入Excel文件
def To_Excel():
    wb=openpyxl.load_workbook(path)
    ws=wb[sheetname[0]]

    ws.cell(row=1,column=4).value='性别'
    for i in range(1,len(sex)+1):
        re=sex[i-1]
        ws.cell(row=i+1,column=4).value=re

    ws.cell(row=1,column=5).value='生日'
    for i in range(1,len(bday)+1):
        er=bday[i-1]
        ws.cell(row=i+1,column=5).value=er

    ws.cell(row=1,column=6).value='年龄'
    for i in range(1,len(Age)+1):
        er=Age[i-1]
        ws.cell(row=i+1,column=6).value=er

    ws.cell(row=1,column=7).value='是否符合要求'
    for i in range(1,len(result)+1):
        er=result[i-1]
        ws.cell(row=i+1,column=7).value=er

    save=wb.save(path)
    
    if save==None:
        print('文件保存成功！')
    else:
        print('文件保存失败！')

#主程序执行入口        
if __name__=='__main__':
    
    ID=[]
    sex=[]
    bday=[]
    Age=[]
    result=[]
    while True:
        
        os.system('clear')    #清屏操作
        path=input('请输入Excel文件的路径，输入exit退出：')
        path=path.strip()          
        #path='/Users/zhaoyang/Desktop/250.xlsx'
        #打开Excel并提取数据

        if os.path.exists(path)==False and path!='exit':
            print('未找到该文件，请检查文件路径！')
            continue
        elif path=='exit':
            sys.exit()
        else:
    
            data=pd.DataFrame(pd.read_excel(path))
            sheetname=pd.ExcelFile(path)
            sheetname=sheetname.sheet_names
        #list(data.columns)

        for i in data['ID']:
            ID.append(str(i))
        
        SEX()
        birthday()
        age()
        Judge()
        To_Excel()
