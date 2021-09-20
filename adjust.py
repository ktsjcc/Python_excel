import os,sys
import time
import openpyxl
import pandas as pd
import random
"""定义变量以及列表初始化"""
coe=[]
rex=[]
re=[]
part=0
count=0

while True:

	os.system('clear')

	print('\033[1;32;40m============================================================================')
	print('这是一个用来调整价格的脚本，首先要将需要调整的全部价格导入到一个Excel 表格中')
	print('然后调用该脚本即可完成操作。第一列列名要改成total！')                                                
	print('============================================================================')
	print('\033[0m')

	path=input('\033[1;33;40m请输入Excel文件路径,输入exit退出: ')
	print('\033[0m')
	path=path.strip()
	if path=='exit':
		sys.exit()

	elif os.path.exists(path)==False:
		print('\033[1;37;31m未找到该文件，请检查文件路径！\033[1;37;31m')
		print('\033[0m')
		time.sleep(1)
		continue

	target=input("请输入目标价格： ")
	#extracting data from excel
	data=pd.DataFrame(pd.read_excel(path))
	for i in data['total']:
		coe.append(i)
	for i in coe:
		count=i+count

	r=float(target)/count     
	for i in range(len(coe)-1):
		rex.append(r+random.uniform(-1/(10**(round(len(target)/2)+1)),1/(10**(round(len(target)/2)+1))))	

	for i in range(len(coe)-1):
		part=part+coe[i]*rex[i]
	lastrex=(float(target)-part)/coe[-1]	
	rex.append(round(lastrex,3))

	#writing to excel
	wb=openpyxl.load_workbook(path)
	ws=wb['Sheet1']

	ws.cell(row=1,column=2).value='系数'
	for i in range(1,len(rex)+1):
		re=rex[i-1]
		ws.cell(row=i+1,column=2).value=re

	ws.cell(row=1,column=3).value='结果'
	for i in range(1,len(rex)+1):
		er=rex[i-1]*coe[i-1]
		ws.cell(row=i+1,column=3).value=er

	save=wb.save(path)
	if save==None:
		print('\033[1;32;40m文件保存成功!\033[1;32;40m')
		print('\033[0m')
		time.sleep(2)
	else:
		print('\033[1;37;31m文件保存失败！\033[1;37;31m')
		print('\033[0m')
		time.sleep(2)
		
